import os
import openpyxl

def extract_data(file, sheet_name):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column
    header = []
    header_row = 0
    
    # loop through the rows to find the header row
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value in ['s/n', 'name', 'phone number', 'bank', 'account number']:
                header_row = row
                break
        if header_row != 0:
            break
    
    # loop through the header row to get the header
    for col in range(1, max_col + 1):
        header.append(sheet.cell(row=header_row, column=col).value)
    
    data = []
    for row in range(header_row + 1, max_row + 1):
        row_data = {}
        for col, header_cell in enumerate(header):
            row_data[header_cell] = sheet.cell(row=row, column=col + 1).value
        data.append(row_data)
    return data

def write_to_excel(data, file_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Extracted Data'
    
    # write the header to the first row
    for col, header in enumerate(data[0].keys()):
        sheet.cell(row=1, column=col + 1, value=header)
    
    # write the data to the remaining rows
    for row, item in enumerate(data):
        for col, header in enumerate(item.keys()):
            sheet.cell(row=row + 2, column=col + 1, value=item[header])
    
    wb.save(file_path)

folder_path = 'path/to/folder'
output_folder = 'path/to/output/folder'

for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file = os.path.join(folder_path, filename)
        data = extract_data(file, 'Sheet1')
        output_file = os.path.join(output_folder, filename)
        write_to_excel(data, output_file)
