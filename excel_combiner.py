import pandas as pd
import openpyxl
import os
import shutil
from openpyxl import load_workbook
import random as rd
import re
import numpy as np


def find_columns_and_rows(file_path):
    # Read the Excel file
    df = pd.read_excel(file_path, engine='openpyxl')

    # Define the column keywords with their possible variations
    keywords = {
        'name': ['name', 'names'],
        'phone': ['phone', 'phone numbers', 'phone number', 'phone no', 'CONTACTS', 'contact'],
        'bank': ['bank', 'banks', 'bank names', 'bank name'],
        'account': ['account number', 'bank account', 'account numbers', 'accounts', 'bank accounts', 'account', 'account no', 'acct no']
    }

    rows_columns_to_extract = {}

    # Find the columns and starting row of data
    for index, row in df.iterrows():
        current_row_columns = {}
        for col_name, variations in keywords.items():
            if col_name not in rows_columns_to_extract:
                for col_idx, cell_value in enumerate(row):
                    if isinstance(cell_value, str):
                        cell_value_lower = cell_value.lower()
                        if any(variation.lower() in cell_value_lower for variation in variations):
                            current_row_columns[col_name] = (index + 1, index + 1000, col_idx)
                            break

        # If all columns are found in the current row, assign the values to rows_columns_to_extract and break the loop
        if len(current_row_columns) == len(keywords):
            rows_columns_to_extract = current_row_columns
            break

    return rows_columns_to_extract


def move_problematic_file(source_file_path, destination_folder):
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    file_name = os.path.basename(source_file_path)
    destination_file_path = os.path.join(destination_folder, file_name)
    shutil.move(source_file_path, destination_file_path)
    print(f"Moved problematic file {source_file_path} to {destination_file_path}")

def convert_xls_to_xlsx(input_file_path, output_file_path):
    # Read the .xls file
    xls_data = pd.read_excel(input_file_path, engine='xlrd')

    # Write the data to an .xlsx file
    xls_data.to_excel(output_file_path, index=False, engine='openpyxl')
# Replace this with the path to your folder


source_folder = r'C:\Users\user\Downloads\Bundles\OKEMAX'
destination_folder = r'C:\Users\user\Downloads\Bundles\Todo-issues'

def wookbook_converter():
    for file in os.listdir(source_folder):
        if file.endswith('.xls'):
            file_path = os.path.join(source_folder, file)
            output_file_path = os.path.splitext(file_path)[0] + '.xlsx'
            
            # Convert the .xls file to .xlsx
            convert_xls_to_xlsx(file_path, output_file_path)
            move_problematic_file(file_path, destination_folder)
            print(f'Converted {file} to .xlsx format.')

    
def excel_combiner():
    # Replace this with the path to your folder
    source_folder = r'C:\Users\user\Downloads\Bundles\OKEMAX'
    output_file = 'combined_data.xlsx'  # The output file name

    # Define the rows and columns you want to extract from the Excel files
    # Format: (row_start, row_end, column_index)
    # rows_columns_to_extract = {
    #     # Rows 1 to 10 (inclusive) in column 1 (0-indexed)
    #     'NAME ': (0, 1000, 1),
    #     # Rows 1 to 10 (inclusive) in column 2 (0-indexed)
    #     'PHONE ': (0, 1000, 2),
    #     # Rows 1 to 10 (inclusive) in column 3 (0-indexed)
    #     'ACCOUNT NO': (0, 1000, 3),
    #     # Rows 1 to 10 (inclusive) in column 4 (0-indexed)
    #     'NAME OF BANK': (0, 1000, 4),
    # }
    rows_columns_to_extract = {
        # Rows 1 to 10 (inclusive) in column 1 (0-indexed)
        'name': (0, 1000, 1),
        # Rows 1 to 10 (inclusive) in column 2 (0-indexed)
        'phone': (0, 1000, 2),
        # Rows 1 to 10 (inclusive) in column 3 (0-indexed)
        'account': (0, 1000, 3),
        # Rows 1 to 10 (inclusive) in column 4 (0-indexed)
        'bank': (0, 1000, 4),
    }

    # Initialize an empty DataFrame to store the combined data
    combined_data = pd.DataFrame(columns=rows_columns_to_extract.keys())

    for file in os.listdir(source_folder):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            file_path = os.path.join(source_folder, file)
            rows_columns_to_extract = find_columns_and_rows(file_path)
            print(f'For {file}: {rows_columns_to_extract}')

             # Check if the 'name' key exists in the rows_columns_to_extract dictionary
            # if 'name' not in rows_columns_to_extract:
            #     print(f"Could not find the required columns and rows in {file}. Skipping this file.")
            #     continue

            try:
                # Read the Excel file
                df = pd.read_excel(file_path, engine='openpyxl')
            except ValueError as e:
                if "Max value is 14" in str(e):
                    print(f"Font family value error in {file}. Moving problematic file.")
                    move_problematic_file(file_path, destination_folder)
                    continue
                else:
                    raise

            # Check if at least 90% of the values in the 'name' column are strings
            name_col_idx = rows_columns_to_extract['name'][2]
            name_col_values = df.iloc[:, name_col_idx]
            string_count = sum(isinstance(value, str) for value in name_col_values)
            string_percentage = string_count / len(name_col_values) * 100
            
            if string_percentage < 90:
                print(f"Column 'NAME ' in {file} has incorrect data type. Skipping this file.")
                move_problematic_file(file_path, destination_folder)
                continue

            # Extract the data for each column using iloc[]
            extracted_data = pd.DataFrame(columns=rows_columns_to_extract.keys())
            for col_name, (row_start, row_end, col_idx) in rows_columns_to_extract.items():
                if col_idx >= len(df.columns):
                    print(f"Column index out-of-bounds for {col_name} in {file}. Skipping this column.")
                    move_problematic_file(file_path, destination_folder)
                    continue
                extracted_data[col_name] = df.iloc[row_start:row_end + 1, col_idx]

            # Append the extracted data to the combined_data DataFrame
            combined_data = pd.concat([combined_data, extracted_data], ignore_index=True)
            print(f'Data extracted from {file} successfully.')

    # Write the combined data to the output Excel file
    combined_data.to_excel(output_file, index=False, engine='openpyxl')
    print(f'Combined data has been saved to {output_file}')


def remove_empty_value():
    input_file = 'combined_data.xlsx'  # The input file with combined data
    output_file = 'cleaned_combined_data.xlsx'  # The output file with cleaned data

    # Read the combined data from the input file
    combined_data = pd.read_excel(input_file, engine='openpyxl')

    # Remove rows with missing values
    cleaned_data = combined_data.dropna(how='any')

    # Write the cleaned data to the output file
    cleaned_data.to_excel(output_file, index=False, engine='openpyxl')
    print(f'Cleaned data has been saved to {output_file}')

import re

def swap_bank_account_if_needed(row):
    bank_cell = str(row['bank'])
    account_cell = str(row['account'])

    # Remove special characters from bank_cell
    bank_cell = re.sub(r'[^\w\s]', '', bank_cell)

    # Pad with zeroes if bank_cell is digits and its length is less than 10 but greater than 5
    if bank_cell.isdigit() and 5 < len(bank_cell) < 10:
        bank_cell = bank_cell.zfill(10)
        row['bank'] = bank_cell

    # If the bank cell contains an account and the account cell contains a bank, swap them
    if (bank_cell.isdigit() and len(bank_cell) == 10 and any(char.isalpha() for char in account_cell)):
        row['bank'], row['account'] = row['account'], row['bank']
    # If both cells contain only integers or only letters, set the 'bank' cell to None to remove the row later
    elif (account_cell.isdigit() and bank_cell.isdigit()) or (account_cell.isalpha() and bank_cell.isalpha()):
        row['bank'] = None

    return row


def format_and_filter_data():
    input_file = 'cleaned_combined_data.xlsx'  # The input file with cleaned data
    output_file_prefix = 'formatted_combined_data'  # The output file prefix for formatted data

    # Read the cleaned data from the input file
    cleaned_data = pd.read_excel(input_file, engine='openpyxl')
    
    # Process the 'name', 'bank', 'account', and 'phone' columns as before
    # Remove "," with space, remove excessive spaces, and convert 'name' column to uppercase
    for index, value in cleaned_data['name'].items():
        if not isinstance(value, str):
            print(f"Value causing error on name cell: {value}")
            print(f"Index of value causing error: {index}")
            break

    for index, value in cleaned_data['bank'].items():
        if not isinstance(value, str):
            print(f"Value causing error on bank cell: {value}")
            print(f"Index of value causing error: {index}")
            break

    cleaned_data['name'] = cleaned_data['name'].apply(lambda x: x.replace(',', ' ').strip().upper())
    print(f'Name proccessed')

    # Apply the swap and cleaning function
    cleaned_data = cleaned_data.apply(swap_bank_account_if_needed, axis=1)

    # Remove rows where bank is None
    cleaned_data = cleaned_data[cleaned_data['bank'].notnull()]

    # Convert 'bank' column to uppercase
    cleaned_data['bank'] = cleaned_data['bank'].astype(str).apply((lambda x: x.replace(',', '').replace("'", '').replace(';', '').strip()))
    cleaned_data = cleaned_data[cleaned_data['bank'].notnull()]
    cleaned_data['bank'] = cleaned_data['bank'].apply(lambda x: x.upper())
    
    print(f'Bank proccessed')

    # Filter out 'account' rows that do not contain any digit
    cleaned_data = cleaned_data[cleaned_data['account'].apply(lambda x: any(char.isdigit() for char in str(x)))]
    
    # Remove "," and spaces, and ensure 'account' column values have exactly 10 digits
    cleaned_data['account'] = cleaned_data['account'].astype(str).apply(lambda x: x.replace(',', '').replace(' ', '').replace('-','').zfill(10))
    print(f'Account proccessed')

    # Process 'phone' column
    cleaned_data['phone'] = cleaned_data['phone'].astype(str).apply(lambda x: x.replace(',', '').replace(' ', '').replace('-',''))
    cleaned_data['phone'] = cleaned_data['phone'].astype(str).apply(lambda x: '0' + x[1:] if x.startswith('+234') else ('0' + x[3:] if x.startswith('234') else x))
    cleaned_data['phone'] = cleaned_data['phone'].astype(str).apply(lambda x: '0' + x if len(x) == 10 and not x.startswith('0') else x)

    # Filter out phone numbers that are not integers
    cleaned_data = cleaned_data[cleaned_data['phone'].apply(lambda x: x.isdigit())]

    # Filter out rows with less than 11 digits in the 'phone' column
    cleaned_data = cleaned_data[cleaned_data['phone'].apply(lambda x: len(x) == 11)]
    print(f'Phone proccessed')

    # Remove rows with missing values
    cleaned_data = cleaned_data.dropna(how='any')
    print(f'Missing values removed')

    # Save the updated output file(s)
    max_rows_per_file = 1000000
    num_output_files = (len(cleaned_data) // max_rows_per_file) + 1

    for file_num in range(num_output_files):
        start_row = file_num * max_rows_per_file
        end_row = start_row + max_rows_per_file
        output_data = cleaned_data.iloc[start_row:end_row]

        output_file = f'{output_file_prefix}_{file_num + 1}.xlsx'

        # Check if the output file exists, if not, create it with headers
        if not os.path.exists(output_file):
            headers = pd.DataFrame(columns=output_data.columns)
            headers.to_excel(output_file, index=False, engine='openpyxl')

        # Load the output file and find the last row with data
        output_workbook = openpyxl.load_workbook(output_file)
        output_sheet = output_workbook.active
        last_row = output_sheet.max_row

        # Append the formatted data to the output file starting from the next row
        for row_index, (index, row) in enumerate(output_data.iterrows(), start=1):
            for col_num, value in enumerate(row, start=1):
                output_sheet.cell(row=last_row + row_index, column=col_num, value=value)

        # Save the output file
        output_workbook.save(output_file)
        print(f'Formatted data has been saved to {output_file}')


output_folder = r'C:\Users\user\Documents\python output\noti'


def write_unique_names(iterations, output_folder):
    # Load the formatted_combined_data DataFrame
    df = pd.read_excel('formatted_combined_data_1.xlsx', engine='openpyxl')
    

    for i in range(iterations):
         # Remove rows that have duplicate 'name', 'phone', and 'account' values
        df = df.drop_duplicates(subset=['name'])
        df = df.drop_duplicates(subset=['phone'])
        df = df.drop_duplicates(subset=['account'])

        # Randomly select 1000 unique rows
        unique_rows = df.sample(n=1000)
        
        # Define the Excel file and the sheet you want to write to
        excel_file = os.path.join(output_folder, f'{i+1}-NOTI-UAAG.xlsx')
        sheet_name = 'Sheet1'

        # Load the format workbook and select the sheet
        format = r'C:\Users\user\Documents\excel formats\NONTI -  UAAG.xlsx'
        book = load_workbook(format)
        sheet = book[sheet_name]
        

        # Define the starting row and column for each column
        start_row = 12  # Start from the second row to avoid overwriting the headers
        name_col = 'B'
        phone_col = 'C'
        account_col = 'E'
        bank_col = 'D'
        sex_col = 'G'

        # Write the data to the specified cells in the Excel file
            # Write the data to the specified cells in the Excel file
        for index, row in unique_rows.iterrows():
            sheet[name_col + str(start_row)] = row['name']
            sheet[phone_col + str(start_row)] ='0' + str(row['phone'])  # prepend single quote to treat as text
            sheet[account_col + str(start_row)] = row['account']
            sheet[bank_col + str(start_row)] =row['bank']
            sheet[sex_col + str(start_row)] = rd.choice(['F','M'])
            start_row += 1

        # Save the changes to the Excel file
        book.save(excel_file)
        print(f'File saved: {excel_file}')




# Run Functions
if __name__ == "__main__":
    # Call the functions
    # wookbook_converter()
    excel_combiner()
    remove_empty_value()
    # format_and_filter_data()
    # write_unique_names(300, output_folder)