import pandas as pd
import os
import shutil
from openpyxl import load_workbook

def find_columns_and_rows(input_file, output_file):
    # Read the Excel files
    df_in = pd.read_excel(input_file, engine='openpyxl')
    df_out = pd.read_excel(output_file, engine='openpyxl')

    # Define the column keywords with their possible variations
    keywords = {
        'name': ['name', 'names','account name'],
        'phone': ['phone', 'phone numbers', 'phone number', 'phone no'],
        'bank': ['bank', 'banks', 'bank names', 'bank name'],
        'account': ['account number', 'bank account', 'account numbers', 'accounts', 'bank accounts', 'account', 'account no']
    }

    output_keywords = {
        'name': ['name', 'names'],
        'phone': ['phone number', 'phone numbers', 'phone no'],
        'bank': ['bank'],
        'account': ['account number', 'account numbers', 'account no']
    }

    rows_columns_to_extract = {}
    output_columns = {}

    # Find the columns and starting row of data in the input file
    for index, row in df_in.iterrows():
        current_row_columns = {}
        for col_name, variations in keywords.items():
            if col_name not in rows_columns_to_extract:
                for col_idx, cell_value in enumerate(row):
                    if isinstance(cell_value, str):
                        cell_value_lower = cell_value.lower()
                        if any(variation.lower() in cell_value_lower for variation in variations):
                            current_row_columns[col_name] = (index + 1, index + 1002, col_idx)
                            break

        # If all columns are found in the current row, assign the values to rows_columns_to_extract and break the loop
        if len(current_row_columns) == len(keywords):
            rows_columns_to_extract = current_row_columns
            break

    # Find the output columns in the output file
    for index, row in df_out.iterrows():
        current_row_columns = {}
        for col_name, variations in output_keywords.items():
            if col_name not in output_columns:
                for col_idx, cell_value in enumerate(row):
                    if isinstance(cell_value, str):
                        cell_value_lower = cell_value.lower()
                        if any(variation.lower() in cell_value_lower for variation in variations):
                            current_row_columns[col_name] = (index + 1, col_idx)
                            break

        # If all columns are found in the current row, assign the values to output_columns and break the loop
        if len(current_row_columns) == len(output_keywords):
            output_columns = current_row_columns
            break

    return rows_columns_to_extract, output_columns



def excel_data_transfer(input_folder, output_folder, template_file):
    

    for file in os.listdir(input_folder):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            file_path = os.path.join(input_folder, file)

            rows_columns_to_extract, output_columns = find_columns_and_rows(file_path, template_file)
            print(f'For {file}: {rows_columns_to_extract}, {output_columns}')
            if not rows_columns_to_extract or not output_columns:
                print(f"Could not find the required columns and rows in {os.path.basename(file)}. Skipping this file.")
                continue

            # Read the data from the input file
            df_in = pd.read_excel(file_path, engine='openpyxl', header=None)

            # Load existing workbook
            book = load_workbook(template_file)

            # Select worksheet
            writer_sheets = {ws.title: ws for ws in book.worksheets}

            for col_name, (start_row, end_row, col_idx) in rows_columns_to_extract.items():
                # Get the existing ExcelWriter sheet
                sheet = writer_sheets['Sheet1']

                # Extract the data from the input DataFrame
                # Reset the index before iterating
                data_to_transfer =  df_in.iloc[start_row+1:end_row, col_idx]
                count = 2
                for i, data in data_to_transfer.iteritems():

                    sheet.cell(row=output_columns[col_name][0] + count , column=output_columns[col_name][1] + 1, value=data)
                    count += 1

            # Save the workbook
            book.save(f'{output_folder}/{os.path.basename(file)}')
            print(f"Transfer data to {file} completEd")



if __name__ == "__main__":
    input_folder = r'C:\Users\user\Documents\python output\Todo Work'
    output_folder = r'C:\Users\user\Documents\python output\Transfer Done'
    template_file = r'C:\Users\user\Documents\excel formats\KOKOOLA MPCSL.xlsx'
    excel_data_transfer(input_folder, output_folder, template_file)