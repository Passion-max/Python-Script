import os
from openpyxl import load_workbook
import openpyxl




def split_worksheets(input_file_path, output_folder, name_cell):
    # Load the input workbook
    count=28
    input_wb = load_workbook(input_file_path)
    print('Workbook loaded...')
    # Iterate through each worksheet
    for ws_name in input_wb.sheetnames:
        print(ws_name + ' Selected...')
        # Create a new workbook for the current worksheet
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active

        # Copy the content of the current worksheet to the new workbook
        input_ws = input_wb[ws_name]
        print('Copying...')
        for row in input_ws.iter_rows():
            for cell in row:
                output_ws[cell.coordinate].value = cell.value

        # Get the new file name from the name_cell
        file_name = "BENJONSYL " + str(count)

        # Save the new workbook
        output_wb.save(f"{output_folder}/{file_name}.xlsx")
        count += 1

if __name__ == "__main__":
    # folder_path = r"C:\Users\user\Downloads\OPE"
    # target_file_path = r"C:\Users\user\Downloads\format\BENSONLY.xlsx"
    # copy_data_to_file(folder_path, target_file_path)

    # # Usage example
    input_file_path = r"C:\Users\user\Downloads\UN-UAAG TEMPLATE VICTORIOUS CHRIST\UN-UAAG TEMPLATE VICTORIOUS CHRIST MAIN.xlsx"
    output_folder = r"C:\Users\user\Downloads\UN-UAAG TEMPLATE VICTORIOUS CHRIST"
    name_cell = "A1"  # Replace this with the cell address containing the desired file name

    split_worksheets(input_file_path, output_folder, name_cell)